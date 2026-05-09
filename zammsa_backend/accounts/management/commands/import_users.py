import csv
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from accounts.models import User
from accounts.serializers import UserCreateSerializer


class Command(BaseCommand):
    help = 'Import users from CSV or Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to CSV or Excel file')
        parser.add_argument('--format', type=str, choices=['csv', 'xlsx'], help='File format')

    def handle(self, *args, **options):
        file_path = options['file_path']
        fmt = options.get('format')

        if not fmt:
            if file_path.endswith('.csv'):
                fmt = 'csv'
            elif file_path.endswith('.xlsx'):
                fmt = 'xlsx'
            else:
                raise CommandError('Could not determine file format. Use --format csv or --format xlsx')

        if fmt == 'csv':
            with open(file_path, 'r') as f:
                reader = csv.DictReader(f)
                self._import_rows(list(reader))
        else:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            self._import_rows(rows)

    def _import_rows(self, rows):
        success = 0
        errors = []
        for i, row in enumerate(rows, start=2):
            serializer = UserCreateSerializer(data=row)
            if serializer.is_valid():
                serializer.save()
                success += 1
            else:
                errors.append(f'Row {i}: {serializer.errors}')
        self.stdout.write(self.style.SUCCESS(f'Imported {success} users'))
        for error in errors:
            self.stdout.write(self.style.ERROR(error))
